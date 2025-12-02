"""
Output Formatters
Formats test case output in multiple formats (JSON, Markdown, Excel)
"""
import sys
import json
import re
from pathlib import Path
from typing import Dict, List
from datetime import datetime

# Add project root to path
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

import config
from src.utils.logger import setup_logger

logger = setup_logger(__name__)


class TestCaseFormatter:
    """Format test cases in multiple output formats"""

    def __init__(self):
        """Initialize formatter"""
        self.output_dir = config.DATA_DIR / "test_cases"
        self.output_dir.mkdir(parents=True, exist_ok=True)
        logger.info(f"Output directory: {self.output_dir}")

    def parse_test_cases(self, test_cases_text: str) -> List[Dict]:
        """
        Parse test cases from text format to structured data

        Args:
            test_cases_text: Generated test cases in text format

        Returns:
            List of test case dictionaries
        """
        logger.info("Parsing test cases from text...")

        # Split by test case ID pattern - match both **TC_XXX** and #### TC_XXX formats
        # GPT-5 uses markdown headings, GPT-4 uses bold
        pattern_bold = r'\*\*TC_\d+\*\*'  # **TC_001**
        pattern_heading = r'####\s*TC_\d+'  # #### TC_001

        # Try both patterns and use whichever finds matches
        test_ids_bold = re.findall(pattern_bold, test_cases_text)
        test_ids_heading = re.findall(pattern_heading, test_cases_text)

        if test_ids_heading:  # GPT-5 format (markdown headings)
            pattern = pattern_heading
            sections = re.split(pattern, test_cases_text)
            test_ids = [tid.replace('####', '').strip() for tid in test_ids_heading]
        else:  # GPT-4 format (bold)
            pattern = pattern_bold
            sections = re.split(pattern, test_cases_text)
            test_ids = [tid.replace('**', '').strip() for tid in test_ids_bold]

        test_cases = []

        for idx, section in enumerate(sections[1:]):  # Skip first section before first TC
            if not section.strip():
                continue

            # Get the test ID for this section
            test_id = test_ids[idx] if idx < len(test_ids) else f"TC_{idx+1:03d}"

            test_case = self._parse_single_test_case(section, test_id)
            if test_case:
                test_cases.append(test_case)

        logger.info(f"Parsed {len(test_cases)} test cases")
        return test_cases

    def _parse_single_test_case(self, text: str, test_id: str = '') -> Dict:
        """Parse a single test case from text"""
        test_case = {
            'test_id': test_id,
            'title': '',
            'category': '',
            'priority': '',
            'description': '',
            'prerequisites': '',
            'test_data': '',
            'test_steps': [],
            'expected_results': '',
            'postconditions': ''
        }

        lines = text.strip().split('\n')

        current_field = None
        for line in lines:
            line = line.strip()

            # Skip empty lines and separators
            if not line or line.startswith('---') or line.startswith('###') or line.startswith('END OF SECTION'):
                continue

            # Match **Field:** format using simple string matching
            if line.startswith('**Test Title:**') or line.startswith('**Title:**'):
                test_case['title'] = line.replace('**Test Title:**', '').replace('**Title:**', '').strip()
                current_field = None
            elif line.startswith('**Category:**'):
                test_case['category'] = line.replace('**Category:**', '').strip()
                current_field = None
            elif line.startswith('**Priority:**'):
                test_case['priority'] = line.replace('**Priority:**', '').strip()
                current_field = None
            elif line.startswith('**Description:**'):
                current_field = 'description'
                test_case['description'] = line.replace('**Description:**', '').strip()
            elif line.startswith('**Prerequisites:**') or line.startswith('**Prerequisite:**'):
                current_field = 'prerequisites'
                test_case['prerequisites'] = line.replace('**Prerequisites:**', '').replace('**Prerequisite:**', '').strip()
            elif line.startswith('**Test Data:**'):
                current_field = 'test_data'
                test_case['test_data'] = line.replace('**Test Data:**', '').strip()
            elif line.startswith('**Test Steps:**'):
                current_field = 'test_steps'
            elif line.startswith('**Expected Results:**') or line.startswith('**Expected Result:**'):
                current_field = 'expected_results'
                test_case['expected_results'] = line.replace('**Expected Results:**', '').replace('**Expected Result:**', '').strip()
            elif line.startswith('**Postconditions:**') or line.startswith('**Postcondition:**'):
                current_field = 'postconditions'
                test_case['postconditions'] = line.replace('**Postconditions:**', '').replace('**Postcondition:**', '').strip()
            elif line and current_field:
                # Append to current field
                if current_field == 'test_steps':
                    # Remove numbering like "1. ", "2. ", etc.
                    cleaned_line = re.sub(r'^\d+\.\s*', '', line)
                    if cleaned_line:
                        test_case['test_steps'].append(cleaned_line)
                else:
                    test_case[current_field] += ' ' + line

        # Clean up fields
        for key in test_case:
            if isinstance(test_case[key], str):
                test_case[key] = test_case[key].strip()

        return test_case if test_case['title'] else None

    def save_as_json(self, result: Dict, filename: str = None) -> str:
        """
        Save test cases as JSON

        Args:
            result: Test generation result
            filename: Optional filename

        Returns:
            Path to saved file
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"test_cases_{timestamp}.json"

        filepath = self.output_dir / filename

        # Parse test cases
        test_cases = self.parse_test_cases(result['test_cases'])

        # Create JSON structure
        output = {
            'generated_at': datetime.now().isoformat(),
            'test_plan': result.get('test_plan', ''),
            'validation_report': result.get('validation_report', ''),
            'total_test_cases': len(test_cases),
            'test_cases': test_cases
        }

        # Save
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(output, f, indent=2, ensure_ascii=False)

        logger.info(f"Saved JSON: {filepath}")
        return str(filepath)

    def save_as_markdown(self, result: Dict, filename: str = None) -> str:
        """
        Save test cases as Markdown

        Args:
            result: Test generation result
            filename: Optional filename

        Returns:
            Path to saved file
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"test_cases_{timestamp}.md"

        filepath = self.output_dir / filename

        # Build markdown
        md_lines = [
            "# Test Case Generation Report",
            f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            "",
            "## Test Planning Analysis",
            "",
            result.get('test_plan', ''),
            "",
            "## Generated Test Cases",
            ""
        ]

        # Parse and format test cases
        test_cases = self.parse_test_cases(result['test_cases'])

        for idx, tc in enumerate(test_cases, 1):
            md_lines.extend([
                f"### Test Case {idx}: {tc['title']}",
                "",
                f"**Category:** {tc['category']}",
                f"**Priority:** {tc['priority']}",
                "",
                f"**Description:** {tc['description']}",
                "",
                f"**Prerequisites:** {tc['prerequisites']}",
                "",
                f"**Test Data:** {tc['test_data']}",
                "",
                "**Test Steps:**",
            ])

            for step in tc['test_steps']:
                md_lines.append(f"- {step}")

            md_lines.extend([
                "",
                f"**Expected Results:** {tc['expected_results']}",
                "",
                f"**Postconditions:** {tc['postconditions']}",
                "",
                "---",
                ""
            ])

        # Add validation report
        md_lines.extend([
            "## Validation Report",
            "",
            result.get('validation_report', ''),
            ""
        ])

        # Save
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write('\n'.join(md_lines))

        logger.info(f"Saved Markdown: {filepath}")
        return str(filepath)

    def save_as_excel(self, result: Dict, filename: str = None) -> str:
        """
        Save test cases as Excel

        Args:
            result: Test generation result
            filename: Optional filename

        Returns:
            Path to saved file
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            logger.error("openpyxl not installed, cannot create Excel file")
            return ""

        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"test_cases_{timestamp}.xlsx"

        filepath = self.output_dir / filename

        # Parse test cases
        test_cases = self.parse_test_cases(result['test_cases'])

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"

        # Header row
        headers = ['Test ID', 'Title', 'Category', 'Priority', 'Description',
                   'Prerequisites', 'Test Data', 'Test Steps', 'Expected Results', 'Postconditions']

        # Style header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Data rows
        for row_idx, tc in enumerate(test_cases, 2):
            test_id = f"TC_{row_idx-1:03d}"
            steps = '\n'.join(tc['test_steps'])

            ws.cell(row=row_idx, column=1, value=test_id)
            ws.cell(row=row_idx, column=2, value=tc['title'])
            ws.cell(row=row_idx, column=3, value=tc['category'])
            ws.cell(row=row_idx, column=4, value=tc['priority'])
            ws.cell(row=row_idx, column=5, value=tc['description'])
            ws.cell(row=row_idx, column=6, value=tc['prerequisites'])
            ws.cell(row=row_idx, column=7, value=tc['test_data'])
            ws.cell(row=row_idx, column=8, value=steps)
            ws.cell(row=row_idx, column=9, value=tc['expected_results'])
            ws.cell(row=row_idx, column=10, value=tc['postconditions'])

            # Wrap text
            for col_idx in range(1, 11):
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True, vertical='top')

        # Adjust column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 50
        ws.column_dimensions['I'].width = 40
        ws.column_dimensions['J'].width = 30

        # Save
        wb.save(filepath)

        logger.info(f"Saved Excel: {filepath}")
        return str(filepath)

    def save_all_formats(self, result: Dict) -> Dict[str, str]:
        """
        Save test cases in all supported formats

        Args:
            result: Test generation result

        Returns:
            Dictionary mapping format to filepath
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        files = {
            'json': self.save_as_json(result, f"test_cases_{timestamp}.json"),
            'markdown': self.save_as_markdown(result, f"test_cases_{timestamp}.md"),
            'excel': self.save_as_excel(result, f"test_cases_{timestamp}.xlsx")
        }

        logger.info(f"Saved test cases in {len(files)} formats")
        return files
